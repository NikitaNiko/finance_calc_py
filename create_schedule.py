import openpyxl
import string
import calendar
from datetime import date, timedelta

import openpyxl.styles

# n -                       количество месяцев выплат
# prcents_and_credit -      сколько нужно отдать банку
# r -                       месячная процентня ставка
# monthly_pay -             ежемесячный платеж
# directody -               директория, куда будет сохраняться файл

def creating(n, percents_and_kredit, r, monthly_pay, directory):
    texts = ["Номер платежа", "Дата платежа", "Остаток долга", "В погашении кредита", "В погашении процентов", "Платеж"]
    book = openpyxl.Workbook()
    sheet = book.active


    for i in range(1, 7):
        sheet.cell(row=1, column=i).value = texts[i - 1]
        sheet.cell(row=1, column=i).font = openpyxl.styles.Font(color="fc0303")
        sheet.column_dimensions[string.ascii_uppercase[i - 1]].width = 20

    today = date.today()
    
    for i in range(2, n + 2):
        sheet.cell(row=i, column=1).value = i - 1                                           # Номера платежей
        sheet.cell(row=i, column=2).value = today 
        today += timedelta(days = calendar.monthrange(today.year, today.month)[1])          # Прибавляем месяц

        percents_and_kredit = round(percents_and_kredit - monthly_pay, 2)
        sheet.cell(row=i, column=3).value = percents_and_kredit
        sheet.cell(row=i, column=4).value = round(monthly_pay - (percents_and_kredit * r), 2)
        sheet.cell(row=i, column=5).value = round(percents_and_kredit * r, 2)
        sheet.cell(row=i, column=6).value = monthly_pay
        

        

    book.save(str(directory))
    book.close()


